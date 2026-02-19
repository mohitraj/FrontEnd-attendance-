from flask import Flask, request, send_file, render_template_string, jsonify
import os
import io
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Attendance Processor</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .card {
            background: white;
            border-radius: 16px;
            padding: 40px;
            width: 100%;
            max-width: 520px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.2);
        }
        h1 {
            font-size: 1.6rem;
            color: #2d3748;
            margin-bottom: 8px;
        }
        p.subtitle {
            color: #718096;
            font-size: 0.9rem;
            margin-bottom: 30px;
        }
        .upload-area {
            border: 2px dashed #cbd5e0;
            border-radius: 12px;
            padding: 40px 20px;
            text-align: center;
            cursor: pointer;
            transition: all 0.2s;
            margin-bottom: 20px;
            position: relative;
        }
        .upload-area:hover, .upload-area.dragover {
            border-color: #667eea;
            background: #f0f4ff;
        }
        .upload-icon { font-size: 2.5rem; margin-bottom: 10px; }
        .upload-area p { color: #718096; font-size: 0.9rem; }
        .upload-area strong { color: #4a5568; }
        #file-input { 
            position: absolute; top: 0; left: 0; 
            width: 100%; height: 100%; 
            opacity: 0; cursor: pointer; 
        }
        #file-name {
            background: #f7fafc;
            border-radius: 8px;
            padding: 10px 15px;
            font-size: 0.85rem;
            color: #4a5568;
            margin-bottom: 20px;
            display: none;
            word-break: break-all;
        }
        button {
            width: 100%;
            padding: 14px;
            background: linear-gradient(135deg, #667eea, #764ba2);
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 1rem;
            font-weight: 600;
            cursor: pointer;
            transition: opacity 0.2s;
        }
        button:hover { opacity: 0.9; }
        button:disabled { opacity: 0.5; cursor: not-allowed; }
        .status {
            margin-top: 20px;
            padding: 12px 16px;
            border-radius: 8px;
            font-size: 0.9rem;
            display: none;
        }
        .status.error { background: #fff5f5; color: #c53030; border: 1px solid #fed7d7; }
        .status.success { background: #f0fff4; color: #276749; border: 1px solid #c6f6d5; }
        .status.loading { background: #ebf8ff; color: #2b6cb0; border: 1px solid #bee3f8; }
        .spinner {
            display: inline-block;
            width: 14px; height: 14px;
            border: 2px solid #2b6cb0;
            border-top-color: transparent;
            border-radius: 50%;
            animation: spin 0.8s linear infinite;
            margin-right: 8px;
            vertical-align: middle;
        }
        @keyframes spin { to { transform: rotate(360deg); } }
        .info-list {
            margin-top: 24px;
            padding-top: 20px;
            border-top: 1px solid #e2e8f0;
        }
        .info-list h3 { color: #4a5568; font-size: 0.85rem; margin-bottom: 10px; }
        .info-list ul {
            list-style: none;
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
        }
        .info-list li {
            background: #edf2f7;
            border-radius: 6px;
            padding: 4px 10px;
            font-size: 0.8rem;
            color: #4a5568;
        }
    </style>
</head>
<body>
<div class="card">
    <h1>📊 Attendance Processor</h1>
    <p class="subtitle">Upload your Excel file to merge lecture &amp; lab attendance into a single formatted sheet.</p>

    <div class="upload-area" id="drop-zone">
        <input type="file" id="file-input" accept=".xlsx">
        <div class="upload-icon">📁</div>
        <p><strong>Click to browse</strong> or drag &amp; drop</p>
        <p>Supports .xlsx files only</p>
    </div>

    <div id="file-name"></div>

    <button id="process-btn" disabled onclick="processFile()">Process Attendance</button>

    <div class="status" id="status"></div>

    <div class="info-list">
        <h3>What this tool does:</h3>
        <ul>
            <li>✅ Merges lecture + lab sheets</li>
            <li>✅ Sorts columns by date &amp; time</li>
            <li>✅ Numbers present attendance</li>
            <li>✅ Marks absences as X (red)</li>
            <li>✅ Highlights lab cols in blue</li>
            <li>✅ Supports sections A–H</li>
        </ul>
    </div>
</div>

<script>
    const fileInput = document.getElementById('file-input');
    const dropZone = document.getElementById('drop-zone');
    const fileNameDiv = document.getElementById('file-name');
    const processBtn = document.getElementById('process-btn');
    const status = document.getElementById('status');
    let selectedFile = null;

    fileInput.addEventListener('change', () => {
        if (fileInput.files[0]) selectFile(fileInput.files[0]);
    });

    dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('dragover'); });
    dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
    dropZone.addEventListener('drop', e => {
        e.preventDefault();
        dropZone.classList.remove('dragover');
        if (e.dataTransfer.files[0]) selectFile(e.dataTransfer.files[0]);
    });

    function selectFile(file) {
        if (!file.name.endsWith('.xlsx')) {
            showStatus('error', 'Please upload a valid .xlsx file.');
            return;
        }
        selectedFile = file;
        fileNameDiv.textContent = '📄 ' + file.name;
        fileNameDiv.style.display = 'block';
        processBtn.disabled = false;
        showStatus('', '');
    }

    function showStatus(type, message) {
        status.className = 'status';
        status.style.display = message ? 'block' : 'none';
        if (type) status.classList.add(type);
        status.innerHTML = message;
    }

    async function processFile() {
        if (!selectedFile) return;
        processBtn.disabled = true;
        showStatus('loading', '<span class="spinner"></span> Processing your file, please wait...');

        const formData = new FormData();
        formData.append('file', selectedFile);

        try {
            const response = await fetch('/process', { method: 'POST', body: formData });
            if (response.ok) {
                const blob = await response.blob();
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = 'final_Attendance.xlsx';
                a.click();
                URL.revokeObjectURL(url);
                showStatus('success', '✅ File processed successfully! Your download should start automatically.');
            } else {
                const err = await response.json();
                showStatus('error', '❌ Error: ' + (err.error || 'Something went wrong.'));
            }
        } catch (e) {
            showStatus('error', '❌ Network error: ' + e.message);
        } finally {
            processBtn.disabled = false;
        }
    }
</script>
</body>
</html>
"""


def process_attendance(input_stream):
    wb = load_workbook(input_stream)
    sections = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    output_wb = Workbook()
    output_wb.remove(output_wb.active)

    sheets_created = 0

    for section in sections:
        if section not in wb.sheetnames:
            continue

        for subsection_num in ['1', '2']:
            subsection_name = f"{section}{subsection_num}"

            if subsection_name not in wb.sheetnames:
                continue

            lecture_sheet = wb[section]
            lab_sheet = wb[subsection_name]

            lecture_start_col = 7 if section in ['C', 'D', 'F', 'G', 'H'] else 8
            lab_start_col = 7 if section in ['C', 'D', 'F', 'G', 'H'] else 8

            def extract_columns(sheet, start_col, source_type):
                columns = []
                for col_idx in range(start_col, sheet.max_column + 1):
                    date_cell = sheet.cell(1, col_idx).value
                    if isinstance(date_cell, datetime):
                        day_cell = sheet.cell(2, col_idx).value
                        if isinstance(day_cell, str) and day_cell.startswith('='):
                            day_cell = date_cell.strftime('%A')
                        time_cell = sheet.cell(3, col_idx).value
                        time_sort_key = 0
                        if time_cell and isinstance(time_cell, str):
                            match = re.search(r'(\d+):(\d+)', time_cell)
                            if match:
                                time_sort_key = int(match.group(1)) * 60 + int(match.group(2))
                        columns.append((date_cell, day_cell, time_cell, col_idx, source_type, time_sort_key))
                return columns

            lecture_columns = extract_columns(lecture_sheet, lecture_start_col, 'lecture')
            lab_columns = extract_columns(lab_sheet, lab_start_col, 'lab')

            all_columns = lecture_columns + lab_columns
            all_columns.sort(key=lambda x: (x[0], x[5]))

            students = []
            for row_idx in range(5, lab_sheet.max_row + 1):
                roll_no = lab_sheet.cell(row_idx, 2).value
                subsection_cell = lab_sheet.cell(row_idx, 5).value

                if roll_no and subsection_cell == subsection_name:
                    student_info = {
                        'sr_no': lab_sheet.cell(row_idx, 1).value,
                        'roll_no': roll_no,
                        'name': lab_sheet.cell(row_idx, 3).value,
                        'section': lab_sheet.cell(row_idx, 4).value,
                        'subsection': subsection_cell,
                        'email': lab_sheet.cell(row_idx, 6).value if lab_sheet.max_column >= 6 else None,
                        'lecture_row': None,
                        'lab_row': row_idx
                    }

                    for lec_row_idx in range(5, lecture_sheet.max_row + 1):
                        lec_roll = lecture_sheet.cell(lec_row_idx, 2).value
                        lec_subsection = lecture_sheet.cell(lec_row_idx, 5).value
                        if lec_roll == roll_no and lec_subsection == subsection_name:
                            student_info['lecture_row'] = lec_row_idx
                            break

                    students.append(student_info)

            new_sheet = output_wb.create_sheet(subsection_name)

            new_sheet['A1'] = 'Sr No'
            new_sheet['B1'] = 'Roll No'
            new_sheet['C1'] = 'Student Name'
            new_sheet['D1'] = 'Section'
            new_sheet['E1'] = 'Subsection'

            has_email = students and students[0].get('email')
            if has_email:
                new_sheet['F1'] = 'Email Id'
                start_att_col = 7
            else:
                start_att_col = 6

            lab_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            absent_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            for idx, (date_val, day_val, time_val, col_idx, source, time_sort) in enumerate(all_columns):
                col_num = start_att_col + idx
                new_sheet.cell(1, col_num).value = date_val
                new_sheet.cell(2, col_num).value = day_val if day_val else date_val.strftime('%A')
                new_sheet.cell(3, col_num).value = time_val if time_val else ''
                if source == 'lab':
                    new_sheet.cell(1, col_num).fill = lab_fill
                    new_sheet.cell(2, col_num).fill = lab_fill
                    new_sheet.cell(3, col_num).fill = lab_fill

            for student_idx, student in enumerate(students):
                row_num = 5 + student_idx
                new_sheet.cell(row_num, 1).value = student['sr_no']
                new_sheet.cell(row_num, 2).value = student['roll_no']
                new_sheet.cell(row_num, 3).value = student['name']
                new_sheet.cell(row_num, 4).value = student['section']
                new_sheet.cell(row_num, 5).value = student['subsection']
                if has_email:
                    new_sheet.cell(row_num, 6).value = student['email']

                present_count = 0
                for idx, (date_val, day_val, time_val, col_idx, source, time_sort) in enumerate(all_columns):
                    col_num = start_att_col + idx
                    attendance_val = None
                    if source == 'lecture' and student['lecture_row']:
                        attendance_val = lecture_sheet.cell(student['lecture_row'], col_idx).value
                    elif source == 'lab' and student['lab_row']:
                        attendance_val = lab_sheet.cell(student['lab_row'], col_idx).value

                    cell = new_sheet.cell(row_num, col_num)
                    if attendance_val == 'P':
                        present_count += 1
                        cell.value = present_count
                        if source == 'lab':
                            cell.fill = lab_fill
                    elif attendance_val == 'A':
                        cell.value = 'X'
                        cell.fill = absent_fill
                        cell.font = Font(color='FFFFFF', bold=True)
                    else:
                        cell.value = ''
                        if source == 'lab':
                            cell.fill = lab_fill

            for cell in new_sheet[1]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

            sheets_created += 1

    if sheets_created == 0:
        raise ValueError(
            "No matching sheets found. Expected sheets like 'A', 'A1', 'A2', 'B', 'B1', etc."
        )

    output_stream = io.BytesIO()
    output_wb.save(output_stream)
    output_stream.seek(0)
    return output_stream


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Please upload a valid .xlsx file'}), 400

    try:
        output_stream = process_attendance(io.BytesIO(file.read()))
        return send_file(
            output_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='final_Attendance.xlsx'
        )
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'Processing failed: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)